# SPDX-License-Identifier: GPL-2.0-or-later
"""Exporting networks and results to CSV and ``.xlsx`` (FR-162).

``specs/17-persistence-and-interoperability.md`` section 5.1.

**The workbook is checked with an independent reader**, not only against the
writer that produced it. A self-consistent round trip through one's own code
proves nothing about whether a spreadsheet application can open the file, which
is the entire point of writing ``.xlsx`` rather than CSV. ``openpyxl`` is a test
dependency for exactly this; the *export* needs nothing but the standard
library, which the "degraded environments" CI job proves by not having it.
"""

from __future__ import annotations

import csv
import zipfile
from xml.etree import ElementTree

import pytest

import tests.networks as nets
from geocomp.core.adjustment import Frame
from geocomp.core.adjustment.least_squares import (
    AdjustmentOptions,
    adjust,
    to_observation_results,
    to_solution,
)
from geocomp.core.errors import ValidationError
from geocomp.core.models import DatumDefinition, HeightType
from geocomp.core.models.epoch import Epoch
from geocomp.core.statistics.tests import data_snooping, global_test
from geocomp.io.tabular import SHEETS, sheet_rows, write_csv, write_workbook


def _openpyxl():
    try:
        import openpyxl as module
    except ImportError:  # pragma: no cover - exercised by the degraded CI job
        return None
    return module


requires_openpyxl = pytest.mark.skipif(
    _openpyxl() is None,
    reason=(
        "openpyxl is a test dependency used to read the workbook back with an "
        "independent implementation; the export itself does not need it"
    ),
)


@pytest.fixture(scope="module")
def adjusted():
    reference = nets.levelling_loop()
    run = adjust(
        reference.network,
        AdjustmentOptions(frame=Frame.HEIGHT_1D, datum=DatumDefinition.CONSTRAINED),
    )
    snooping = data_snooping(
        run.residuals,
        run.cofactor_residuals,
        run.system.weight,
        run.system.row_labels,
        variance_factor=run.variance_factor_aposteriori,
        degrees_of_freedom=run.degrees_of_freedom,
    )
    solution = to_solution(
        run,
        reference.network,
        solution_id="s1",
        crs="EPSG:31982",
        epoch=Epoch.from_decimal_year(2026.0),
        datum=DatumDefinition.CONSTRAINED,
        height_type=HeightType.ORTHOMETRIC,
        observation_results=to_observation_results(run, snooping=snooping),
        global_test=global_test(run.variance_factor_aposteriori, run.degrees_of_freedom),
    )
    return reference.network, solution


class TestTheSheetsAreDeclaredOnce:
    def test_both_writers_share_one_declaration(self):
        """A CSV export and a workbook of the same solution must have the same
        columns in the same order, or one is not a substitute for the other."""
        assert len(SHEETS) == 5
        assert {sheet.name for sheet in SHEETS} == {
            "stations",
            "observations",
            "adjusted",
            "residuals",
            "statistics",
        }

    def test_every_row_matches_its_header_width(self, adjusted):
        network, solution = adjusted
        for sheet in SHEETS:
            headers, rows = sheet_rows(sheet.name, network, solution)
            for row in rows:
                assert len(row) == len(headers), sheet.name

    def test_an_unknown_sheet_is_refused_by_name(self, adjusted):
        network, solution = adjusted
        with pytest.raises(ValidationError) as caught:
            sheet_rows("vibes", network, solution)
        assert caught.value.code == "validation.unknown_export_sheet"

    def test_one_row_per_observation_component(self, adjusted):
        """A GNSS baseline is three rows; merging them would hide which
        component carries the residual."""
        network, solution = adjusted
        _headers, rows = sheet_rows("observations", network, solution)
        expected = sum(len(o.values) for o in network.observations.values())
        assert len(rows) == expected


class TestCsv:
    def test_it_writes_a_file_per_sheet(self, adjusted, tmp_path):
        network, solution = adjusted
        written = write_csv(tmp_path, network=network, solution=solution)
        assert {path.name for path in written} == {
            "stations.csv",
            "observations.csv",
            "adjusted.csv",
            "residuals.csv",
            "statistics.csv",
        }

    def test_an_empty_sheet_is_not_written(self, adjusted, tmp_path):
        """An empty residuals.csv beside an unadjusted network invites the
        reader to conclude the residuals were zero."""
        network, _solution = adjusted
        written = write_csv(tmp_path, network=network)
        assert {path.name for path in written} == {"stations.csv", "observations.csv"}

    def test_values_keep_full_precision(self, adjusted, tmp_path):
        """A coordinate rounded on the way out is rounded for whatever the
        reader does next."""
        network, solution = adjusted
        write_csv(tmp_path, network=network, solution=solution)
        with open(tmp_path / "observations.csv", encoding="utf-8") as handle:
            rows = {row["id"]: row for row in csv.DictReader(handle)}
        for identifier, observation in network.observations.items():
            assert float(rows[identifier]["value"]) == observation.value.value

    def test_the_uncertainty_mode_travels_with_the_value(self, adjusted, tmp_path):
        """FR-203: the distinction must survive to every export."""
        network, solution = adjusted
        write_csv(tmp_path, network=network, solution=solution)
        with open(tmp_path / "observations.csv", encoding="utf-8") as handle:
            rows = list(csv.DictReader(handle))
        assert all(row["uncertainty_mode"] in ("RIGOROUS", "APPROXIMATE") for row in rows)

    def test_a_prefix_is_honoured(self, adjusted, tmp_path):
        network, solution = adjusted
        written = write_csv(tmp_path, network=network, solution=solution, prefix="rd03-")
        assert all(path.name.startswith("rd03-") for path in written)

    def test_selecting_sheets_writes_only_those(self, adjusted, tmp_path):
        network, solution = adjusted
        written = write_csv(
            tmp_path, network=network, solution=solution, sheets=["statistics"]
        )
        assert [path.name for path in written] == ["statistics.csv"]


class TestTheWorkbook:
    def test_it_is_a_zip_with_the_expected_parts(self, adjusted, tmp_path):
        network, solution = adjusted
        path = write_workbook(tmp_path / "x.xlsx", network=network, solution=solution)
        with zipfile.ZipFile(path) as archive:
            names = set(archive.namelist())
        assert "[Content_Types].xml" in names
        assert "xl/workbook.xml" in names
        assert "xl/worksheets/sheet1.xml" in names

    def test_every_part_is_well_formed_xml(self, adjusted, tmp_path):
        network, solution = adjusted
        path = write_workbook(tmp_path / "x.xlsx", network=network, solution=solution)
        with zipfile.ZipFile(path) as archive:
            for name in archive.namelist():
                ElementTree.fromstring(archive.read(name))

    def test_it_is_byte_identical_across_runs(self, adjusted, tmp_path):
        """NFR-007. A file whose bytes change with the clock cannot be
        compared, checksummed or committed."""
        network, solution = adjusted
        first = write_workbook(tmp_path / "a.xlsx", network=network, solution=solution)
        second = write_workbook(tmp_path / "b.xlsx", network=network, solution=solution)
        assert first.read_bytes() == second.read_bytes()

    def test_exporting_nothing_is_refused(self, tmp_path):
        """A workbook of empty sheets says the data was zero rather than absent."""
        with pytest.raises(ValidationError) as caught:
            write_workbook(tmp_path / "x.xlsx")
        assert caught.value.code == "validation.nothing_to_export"

    @requires_openpyxl
    def test_an_independent_reader_opens_it(self, adjusted, tmp_path):
        """The check that matters. A round trip through GeoComp's own writer
        proves nothing about whether a spreadsheet application can open the
        file, which is the entire point of writing .xlsx rather than CSV."""
        module = _openpyxl()
        network, solution = adjusted
        path = write_workbook(tmp_path / "x.xlsx", network=network, solution=solution)

        book = module.load_workbook(path)
        assert book.sheetnames == [
            "stations",
            "observations",
            "adjusted",
            "residuals",
            "statistics",
        ]

    @requires_openpyxl
    def test_the_independent_reader_sees_the_right_headers(self, adjusted, tmp_path):
        module = _openpyxl()
        network, solution = adjusted
        path = write_workbook(tmp_path / "x.xlsx", network=network, solution=solution)
        book = module.load_workbook(path)

        for sheet in SHEETS:
            headers, rows = sheet_rows(sheet.name, network, solution)
            if not rows:
                continue
            worksheet = book[sheet.name]
            found = next(worksheet.iter_rows(max_row=1, values_only=True))
            assert list(found) == list(headers), sheet.name

    @requires_openpyxl
    def test_numbers_arrive_as_numbers_not_text(self, adjusted, tmp_path):
        """A spreadsheet of numbers stored as text is a spreadsheet nobody can
        compute with, and it looks correct until they try."""
        module = _openpyxl()
        network, solution = adjusted
        path = write_workbook(tmp_path / "x.xlsx", network=network, solution=solution)
        worksheet = module.load_workbook(path)["observations"]

        headers = next(worksheet.iter_rows(max_row=1, values_only=True))
        value_column = list(headers).index("value")
        seen = 0
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            assert isinstance(row[value_column], float), row
            seen += 1
        assert seen == len(network.observations)

    @requires_openpyxl
    def test_the_values_survive_the_independent_reader(self, adjusted, tmp_path):
        module = _openpyxl()
        network, solution = adjusted
        path = write_workbook(tmp_path / "x.xlsx", network=network, solution=solution)
        worksheet = module.load_workbook(path)["observations"]

        headers = list(next(worksheet.iter_rows(max_row=1, values_only=True)))
        by_id = {
            row[headers.index("id")]: row for row in worksheet.iter_rows(min_row=2, values_only=True)
        }
        for identifier, observation in network.observations.items():
            assert by_id[identifier][headers.index("value")] == observation.value.value
